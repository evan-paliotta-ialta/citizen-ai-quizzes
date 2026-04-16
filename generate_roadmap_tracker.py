#!/usr/bin/env python3
"""Generate the Citizen AI Engineer Roadmap Task Tracker as an Excel workbook."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

output_path = "/Users/evanpaliotta/Desktop/iAltA Test/Citizen AI Engineer/Citizen AI Engineer Program - Roadmap Tracker.xlsx"

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────────────────────────────
DARK_BLUE  = "1A1A2E"
MID_BLUE   = "2F5496"
LIGHT_BLUE = "D6E4F0"
GREEN      = "C6EFCE"
YELLOW     = "FFEB9C"
ORANGE     = "FFDBB5"
RED_LIGHT  = "FFC7CE"
GRAY_LIGHT = "F5F7FA"
WHITE      = "FFFFFF"

header_font   = Font(name="Calibri", bold=True, size=10, color=WHITE)
header_fill   = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
phase_font    = Font(name="Calibri", bold=True, size=10, color=MID_BLUE)
phase_fill    = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
normal_font   = Font(name="Calibri", size=9.5)
bold_font     = Font(name="Calibri", bold=True, size=9.5)
title_font    = Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
subtitle_font = Font(name="Calibri", size=10, color="888888", italic=True)

wrap_top   = Alignment(wrap_text=True, vertical="top")
wrap_mid   = Alignment(wrap_text=True, vertical="center")
center_mid = Alignment(wrap_text=True, vertical="center", horizontal="center")

thin = Side(style="thin", color="CCCCCC")
med  = Side(style="medium", color="AAAAAA")

def thin_border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def phase_border():
    return Border(left=med, right=med, top=med, bottom=med)

STATUS_COLORS = {
    "Not Started": ("F5F7FA", "555555"),
    "In Progress": (YELLOW,   "9C6500"),
    "Complete":    (GREEN,    "006100"),
    "Blocked":     (RED_LIGHT,"9C0006"),
    "N/A":         ("EEEEEE", "888888"),
}

# ── Tracker sheet ────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Roadmap Tracker"

# Title block
ws.merge_cells("A1:H1")
ws["A1"] = "Citizen AI Engineer Program — Roadmap Tracker"
ws["A1"].font = title_font
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:H2")
ws["A2"] = "Track implementation progress across all phases. Update Status and Date Completed as tasks are finished."
ws["A2"].font = subtitle_font
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

ws.row_dimensions[3].height = 8  # spacer

# Column headers — row 4
headers = ["Phase", "Task #", "Task", "Description", "Owner", "Status", "Date Completed", "Notes"]
col_widths = [18, 8, 36, 52, 16, 14, 16, 30]

for col, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_mid
    cell.border = thin_border()
    ws.column_dimensions[get_column_letter(col)].width = width

ws.row_dimensions[4].height = 20

# Freeze panes below header
ws.freeze_panes = "A5"

# Data validation for Status column (F)
dv = DataValidation(
    type="list",
    formula1='"Not Started,In Progress,Complete,Blocked,N/A"',
    allow_blank=True,
    showDropDown=False,
)
dv.error       = "Please select a valid status."
dv.errorTitle  = "Invalid Status"
dv.prompt      = "Select a status"
dv.promptTitle = "Status"
ws.add_data_validation(dv)
dv.sqref = "F5:F200"

# ── Task data ─────────────────────────────────────────────────────────────────
tasks = [
    # (phase_label, task_num, task, description, owner, status)
    # PHASE 0
    ("Phase 0 — Get the Foundation Approved", None, None, None, None, None),
    ("Phase 0", "0.1", "CIO decision: Enterprise vs Teams plan",
     "Evaluate the compliance and cost analysis. Make the go/no-go decision on Claude Enterprise vs Teams. This gates all subsequent phases.",
     "", "Not Started"),
    ("Phase 0", "0.2", "Enterprise procurement",
     "Purchase Claude Enterprise (minimum 20 seats). Options: self-serve at anthropic.com, sales-assisted, or AWS Marketplace.",
     "", "Not Started"),
    ("Phase 0", "0.3", "Draft Acceptable Use Policy (AUP)",
     "Legal/InfoSec drafts the AI AUP covering Driving Zones, data classification, No-Drive prohibitions, enforcement, and signed acknowledgment requirement.",
     "", "Not Started"),
    ("Phase 0", "0.4", "Update Information Security Policy",
     "Amend the existing InfoSec policy to include AI tools, referencing the Driving Zones model and the AUP.",
     "", "Not Started"),
    ("Phase 0", "0.5", "Brief Board of Directors",
     "Present the Citizen AI Engineer initiative to the Board: scope, compliance posture, governance model, and rollout plan.",
     "", "Not Started"),

    # PHASE 1
    ("Phase 1 — Build the Guardrails", None, None, None, None, None),
    ("Phase 1", "1.1", "Deploy managed-settings.json + managed-mcp.json",
     "Configure and push org-wide managed settings via the Claude Enterprise admin console. Sets permission rules, sandbox enforcement, and MCP deny-by-default.",
     "", "Not Started"),
    ("Phase 1", "1.2", "Write company-wide CLAUDE.md template",
     "Author the global CLAUDE.md: Jira-first rule, no credentials policy, Zone rules, approved tool list, and escalation instructions.",
     "", "Not Started"),
    ("Phase 1", "1.3", "Configure PreToolUse blocking hooks",
     "Write and deploy hooks that block dangerous commands before execution: rm -rf, push to main, secret/credential string patterns.",
     "", "Not Started"),
    ("Phase 1", "1.4", "Configure PostToolUse audit hooks",
     "Write and deploy hooks that log all tool use to the SIEM via the Compliance API. Every Claude action leaves an audit trail.",
     "", "Not Started"),
    ("Phase 1", "1.5", "Set up SCIM provisioning",
     "Configure SCIM between the identity provider and Claude Enterprise. Automates onboarding/offboarding. Directly remediates SOC2 audit exceptions CC5.1.1 and CC6.1.2.",
     "", "Not Started"),
    ("Phase 1", "1.6", "Verify audit log pipeline",
     "Confirm audit logs are flowing correctly from Claude Enterprise to the SIEM. Spot-check a sample of events. Sign off before rollout.",
     "", "Not Started"),

    # PHASE 2
    ("Phase 2 — Build the Certification Program", None, None, None, None, None),
    ("Phase 2", "2.1", "Make course strategy decision (Option A / B / C)",
     "Decide: Anthropic Academy only (A), internal only (B), or hybrid (C). Recommended: hybrid — Academy for Module 1 fundamentals, internal for Modules 2 and 3.",
     "", "Not Started"),
    ("Phase 2", "2.2", "Audit Anthropic Academy catalog",
     "Review current Anthropic Academy course offerings. Designate specific courses for Module 1 (AI Fundamentals). Confirm availability and certificate/proof-of-completion mechanism.",
     "", "Not Started"),
    ("Phase 2", "2.3", "Build Module 2 — Responsible Use at iAltA",
     "Develop the internal course covering: Driving Zones in full (Zones 1–3 + No-Drive), AUP walkthrough, data classification framework, real iAltA scenarios and edge cases.",
     "", "Not Started"),
    ("Phase 2", "2.4", "Build Module 3 — Advanced Workflows",
     "Develop the internal course covering: approved MCP servers and their Zone 3 implications, agentic task design, Jira-first enforcement, CLAUDE.md awareness, hands-on Zone 3 exercise.",
     "", "Not Started"),
    ("Phase 2", "2.5", "Build assessments and quizzes",
     "Write a quiz for each module and a final assessment. Employees must pass the full assessment to receive their license.",
     "", "Not Started"),
    ("Phase 2", "2.6", "Create completion tracker and intake process",
     "Set up the central certification tracker (HubSpot form, Google Sheet, or equivalent). Define the intake process: submission → review → certification record confirmed → access provisioned.",
     "", "Not Started"),
    ("Phase 2", "2.7", "Design the license record",
     "Create the template for the signed, dated completion document issued to each employee upon certification. This is the official record of their license.",
     "", "Not Started"),

    # PHASE 3
    ("Phase 3 — Company-Wide Rollout", None, None, None, None, None),
    ("Phase 3", "3.1", "Confirm all employees complete full certification",
     "Verify that every employee has completed all three modules, passed the assessment, and signed the AUP before any Claude access is provisioned. No exceptions.",
     "", "Not Started"),
    ("Phase 3", "3.2", "Configure and test SCIM-based provisioning",
     "Run end-to-end test of the SCIM provisioning flow: certification confirmed → access granted. Test deprovisioning on a test account.",
     "", "Not Started"),
    ("Phase 3", "3.3", "Company-wide announcement (all-hands)",
     "Announce the program is live. Communicate that using AI without a license is a policy violation. Provide instructions for starting certification.",
     "", "Not Started"),
    ("Phase 3", "3.4", "Certification tracker live and maintained",
     "Confirm the central tracker is operational and being actively maintained. All completions are logged. Access control cross-references the tracker.",
     "", "Not Started"),

    # PHASE 4
    ("Phase 4 — Steady State", None, None, None, None, None),
    ("Phase 4", "4.1", "Select ongoing engagement model",
     "CIO selects the engagement model for ongoing program management: Option 1 (1-on-1 quarterly), Option 2 (weekly office hours), Option 3 (monthly mandatory group), Option 4 (tiered by usage), or Option 5 (async-first).",
     "", "Not Started"),
    ("Phase 4", "4.2", "[Annual] AUP re-certification",
     "All employees re-read the AUP and re-sign annually. Update the certification tracker with renewal dates.",
     "", "Not Started"),
    ("Phase 4", "4.3", "[Annual] Vendor compliance review",
     "Obtain and review the updated Anthropic SOC2 Type II report. Confirm no new compliance gaps. Document findings.",
     "", "Not Started"),
    ("Phase 4", "4.4", "[Annual] Curriculum review",
     "Audit all three course modules for accuracy and relevance. Update content for new Claude capabilities, policy changes, new approved MCP servers, and any zone clarifications.",
     "", "Not Started"),
    ("Phase 4", "4.5", "[Annual] Risk assessment update",
     "Update the organizational AI risk assessment to reflect current usage patterns, incidents, and any changes to the threat landscape or compliance requirements.",
     "", "Not Started"),
    ("Phase 4", "4.6", "[Annual] Incident response drill",
     "Run a tabletop exercise with an AI-specific scenario (e.g., a Zone 3 violation, a credential leak via Claude, or an agentic task that exceeded scope). Document lessons learned.",
     "", "Not Started"),
]

current_row = 5

for item in tasks:
    phase, num, task, desc, owner, status = item
    is_phase_header = num is None

    if is_phase_header:
        # Phase header row
        ws.merge_cells(f"A{current_row}:H{current_row}")
        cell = ws.cell(row=current_row, column=1, value=phase)
        cell.font = phase_font
        cell.fill = phase_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border = phase_border()
        ws.row_dimensions[current_row].height = 18
        current_row += 1
        continue

    row_data = [phase, num, task, desc, owner, status, "", ""]
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=current_row, column=col, value=value)
        cell.font = normal_font
        cell.alignment = wrap_top
        cell.border = thin_border()

        # Alternate row shading
        if (current_row % 2) == 0:
            cell.fill = PatternFill(start_color=GRAY_LIGHT, end_color=GRAY_LIGHT, fill_type="solid")

    # Bold the task name
    ws.cell(row=current_row, column=3).font = bold_font

    # Color-code status cell
    status_cell = ws.cell(row=current_row, column=6)
    bg, fg = STATUS_COLORS.get(status, (GRAY_LIGHT, "555555"))
    status_cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    status_cell.font = Font(name="Calibri", size=9.5, bold=True, color=fg)
    status_cell.alignment = center_mid

    # Center task number
    ws.cell(row=current_row, column=2).alignment = center_mid

    ws.row_dimensions[current_row].height = 42
    current_row += 1

# ── Legend sheet ──────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Legend & Instructions")

ws2["A1"] = "How to Use This Tracker"
ws2["A1"].font = title_font
ws2.row_dimensions[1].height = 24
ws2.merge_cells("A1:D1")

instructions = [
    ("", ""),
    ("COLUMN", "DESCRIPTION"),
    ("Phase", "Which phase of the roadmap this task belongs to."),
    ("Task #", "Unique identifier for cross-referencing with the roadmap document."),
    ("Task", "Short name for the task."),
    ("Description", "Full context for what needs to be done and why."),
    ("Owner", "The person responsible for completing this task. Fill in as assignments are made."),
    ("Status", "Select from the dropdown: Not Started / In Progress / Complete / Blocked / N/A"),
    ("Date Completed", "Enter the date when the task reaches Complete status (format: MM/DD/YYYY)."),
    ("Notes", "Any context, blockers, decisions made, or links to supporting documents."),
    ("", ""),
    ("STATUS COLORS", ""),
    ("Not Started", "Task has not been started yet (default)."),
    ("In Progress", "Task is actively being worked on."),
    ("Complete", "Task is finished and verified."),
    ("Blocked", "Task cannot proceed — note the blocker in the Notes column."),
    ("N/A", "Task is not applicable given decisions made (e.g., task for Option A if Option C was chosen)."),
]

for r, (col_a, col_b) in enumerate(instructions, 2):
    a = ws2.cell(row=r, column=1, value=col_a)
    b = ws2.cell(row=r, column=2, value=col_b)
    a.font = Font(name="Calibri", size=9.5, bold=(col_a in ("COLUMN", "STATUS COLORS")))
    b.font = Font(name="Calibri", size=9.5)
    a.alignment = wrap_top
    b.alignment = wrap_top

    if col_a in ("COLUMN", "STATUS COLORS"):
        a.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        b.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
        a.font = Font(name="Calibri", size=9.5, bold=True, color=MID_BLUE)
        b.font = Font(name="Calibri", size=9.5, bold=True, color=MID_BLUE)

    # Color the status legend rows
    status_colors = {
        "Not Started": STATUS_COLORS["Not Started"],
        "In Progress":  STATUS_COLORS["In Progress"],
        "Complete":     STATUS_COLORS["Complete"],
        "Blocked":      STATUS_COLORS["Blocked"],
        "N/A":          STATUS_COLORS["N/A"],
    }
    if col_a in status_colors:
        bg, fg = status_colors[col_a]
        a.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        a.font = Font(name="Calibri", size=9.5, bold=True, color=fg)

ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 65

wb.save(output_path)
print(f"Roadmap tracker generated: {output_path}")
