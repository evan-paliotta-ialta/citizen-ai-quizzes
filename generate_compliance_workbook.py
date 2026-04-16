import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────────────────────
header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
subheader_font = Font(name="Calibri", bold=True, size=11, color="2F5496")
wrap = Alignment(wrap_text=True, vertical="top")
wrap_center = Alignment(wrap_text=True, vertical="top", horizontal="center")
bold_font = Font(name="Calibri", bold=True, size=11)
normal_font = Font(name="Calibri", size=11)
title_font = Font(name="Calibri", bold=True, size=14, color="2F5496")
subtitle_font = Font(name="Calibri", bold=True, size=12, color="2F5496")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

high_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
high_font = Font(name="Calibri", bold=True, color="9C0006")
med_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
med_font = Font(name="Calibri", bold=True, color="9C6500")
low_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
low_font = Font(name="Calibri", bold=True, color="006100")


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_center
        cell.border = thin_border


def style_row(ws, row, cols, is_section=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.alignment = wrap
        cell.border = thin_border
        cell.font = normal_font
        if is_section:
            cell.fill = subheader_fill
            cell.font = subheader_font


def apply_risk_format(cell, risk):
    if risk == "HIGH":
        cell.fill = high_fill
        cell.font = high_font
    elif risk == "MEDIUM":
        cell.fill = med_fill
        cell.font = med_font
    elif risk == "LOW":
        cell.fill = low_fill
        cell.font = low_font


def auto_width(ws, cols, min_w=12, max_w=45):
    for c in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = max(min_w, min(max_w, 30))


# ════════════════════════════════════════════════════════════════════
# TAB 1 — Control Assessment
# ════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Control Assessment"
ws1.sheet_properties.tabColor = "2F5496"

headers1 = [
    "SOC 2 Ref", "Control", "Requirement Summary",
    "Teams Capability", "Teams Risk", "Risk Rationale (Teams)",
    "Required Compensating Controls (Teams)",
    "Enterprise Capability", "Enterprise Risk", "Risk Rationale (Enterprise)",
    "Audit Exception?"
]
ncols1 = len(headers1)

# Title
ws1.merge_cells("A1:K1")
ws1["A1"].value = "iAltA SOC 2 Type 2 — Control Assessment: Claude Teams vs Enterprise"
ws1["A1"].font = title_font
ws1["A1"].alignment = Alignment(vertical="center")
ws1.row_dimensions[1].height = 30

ws1.merge_cells("A2:K2")
ws1["A2"].value = "Audit Period: July 1 – December 31, 2025 | Auditor: Richey May | System: Verivend | Trust Principle: Security"
ws1["A2"].font = Font(name="Calibri", italic=True, size=10, color="666666")
ws1.row_dimensions[2].height = 20

# Headers in row 4
for i, h in enumerate(headers1, 1):
    ws1.cell(row=4, column=i, value=h)
style_header(ws1, 4, ncols1)

# Column widths
col_widths = [12, 18, 30, 30, 12, 35, 40, 30, 14, 35, 14]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

rows1 = [
    # ── CC5/CC6 Logical & Physical Access Controls ──
    ("section", "CC5 / CC6 — Logical & Physical Access Controls (HIGHEST RISK AREA — Both Existing Exceptions Fall Here)"),
    ("CC5.1.1", "Access Provisioning",
     "Access requests to information resources are documented and approved by Management based on least privilege, need to know, and segregation of duties principles.",
     "SSO + JIT provisioning only. No SCIM for automated lifecycle management. No audit trail of who approved access or when.",
     "HIGH",
     "ALREADY AN AUDIT EXCEPTION. Auditor found no supporting documentation for access requests. Teams lacks SCIM and audit logs, making it impossible to auto-generate the evidence trail the auditor required. JIT provisioning creates accounts on first login but does not document managerial approval.",
     "1) Documented onboarding checklist with manager sign-off ticket (Jira/Drata) for each Claude access grant.\n2) Screenshot or export of admin panel after each provisioning event.\n3) Monthly reconciliation of IdP users vs Claude Teams users.\n4) Store all evidence in Drata or equivalent GRC platform.",
     "SCIM 2.0 auto-syncs provisioning with IdP. Audit logs capture all access grant events with timestamps. Compliance API enables programmatic monitoring.",
     "LOW",
     "SCIM automates the exact workflow the auditor flagged. Audit logs provide the 'supporting documentation' that was missing. Enterprise directly remediates this exception.",
     "YES"),

    ("CC6.1.2", "Access Deprovisioning",
     "System and physical access is revoked within one business day of effective termination date for all terminated users.",
     "No SCIM = manual-only deprovisioning. No audit log to prove timing of revocation. SSO disablement in IdP may not instantly revoke active Claude sessions.",
     "HIGH",
     "ALREADY AN AUDIT EXCEPTION. Auditor found system access was NOT revoked within one business day and access removal ticket documentation was missing. Without SCIM, you rely on a human remembering to manually remove the user from Claude Teams within 24 hours — and documenting it.",
     "1) Add Claude to offboarding checklist with explicit <1 business day SLA.\n2) Automate via IdP: when user is deactivated in Okta/Azure AD, SSO session should be invalidated.\n3) Screenshot Claude admin panel showing user removal with timestamp.\n4) Weekly sweep of terminated employee list vs active Claude users.\n5) Document all in Drata with timestamps.",
     "SCIM 2.0 auto-deprovisions when user is removed from IdP group. Audit logs record exact timestamp of access removal. Gap between termination and revocation is eliminated.",
     "LOW",
     "SCIM directly solves the 1-business-day SLA requirement. Audit logs provide the 'access removal ticket' documentation that was missing. Enterprise directly remediates this exception.",
     "YES"),

    ("CC6.1.3", "Multi-Factor Authentication",
     "Authentication to systems requires the use of multi-factor authentication.",
     "MFA enforced through SSO provider (Okta/Azure AD). Supported on Teams plan.",
     "LOW",
     "Teams supports SSO with SAML 2.0/OIDC. MFA is enforced at the IdP level, which is the industry standard approach. No gap here.",
     "N/A — No compensating controls needed. Ensure MFA is enforced in your IdP configuration.",
     "Same — MFA enforced through SSO/IdP.",
     "LOW",
     "Identical capability. MFA is an IdP-level control, not plan-dependent.",
     "NO"),

    ("CC6.1.4", "Password Policy",
     "Minimum requirements for passwords are documented and enforced for all systems in accordance with Company policy.",
     "Password policy enforced through SSO/IdP. Teams defers to your IdP's password complexity, rotation, and history settings.",
     "LOW",
     "Password policy is controlled at the IdP level (Okta/Azure AD), not within Claude itself. As long as your IdP enforces your documented password policy, this control is satisfied.",
     "N/A — Ensure your IdP password policy matches your documented Information Security Policy.",
     "Same — password policy enforced through SSO/IdP.",
     "LOW",
     "Identical capability. Password policy is IdP-level.",
     "NO"),

    ("CC6.1.5", "Cloud Resources Public Access",
     "Cloud resources are configured to deny public access.",
     "Anthropic manages Claude's infrastructure. Their SOC 2 Type II covers this control. Claude web app requires authentication.",
     "LOW",
     "This is a vendor-side control. Anthropic's own SOC 2 Type II report (available under NDA at trust.anthropic.com) covers their cloud infrastructure security. Claude requires authentication — no anonymous/public access to your workspace.",
     "N/A — Review Anthropic's SOC 2 Type II report annually as part of vendor risk management (CC3.2.2).",
     "Same — Anthropic infrastructure control. Enterprise adds IP allowlisting for additional restriction.",
     "LOW",
     "Same vendor-side control, with the added benefit of IP allowlisting to restrict access to approved corporate networks.",
     "NO"),

    ("CC6.1.6", "Unique Identities",
     "Authentication to systems requires the use of unique identities.",
     "Each Claude Teams user has a unique account tied to their corporate email via SSO.",
     "LOW",
     "SSO integration ensures 1:1 mapping between corporate identity and Claude account. No shared/generic accounts possible when SSO is enforced with domain capture.",
     "N/A — Ensure domain capture is enabled to prevent non-corporate account creation.",
     "Same — unique identities via SSO.",
     "LOW",
     "Identical capability.",
     "NO"),

    ("CC6.1.7", "Privileged Access Restriction",
     "Administrative or privileged access to systems, resources, and functions is restricted to authorized personnel.",
     "Teams has 4 roles: Primary Owner, Owner, Admin, Member. Basic but functional role hierarchy.",
     "MEDIUM",
     "Teams RBAC has only 4 roles with limited granularity. You cannot create custom roles or fine-tune permissions beyond the built-in hierarchy. For a small team (38 users) this is likely sufficient, but audit evidence must clearly document who has each role and why.",
     "1) Document the role assignment rationale for each Admin/Owner.\n2) Include Claude role assignments in annual access review (CC4.1.2).\n3) Minimize Owner/Admin count — principle of least privilege.",
     "Fine-grained RBAC with SCIM-based automated role assignment synced with IdP groups. Managed policy settings for Claude Code (tool permissions, file access, MCP configs).",
     "LOW",
     "Enterprise offers granular, policy-driven RBAC that can be mapped to your org structure and enforced via SCIM group membership. Stronger evidence for auditors.",
     "NO"),

    ("CC4.1.2", "User Access Reviews",
     "Management performs user access reviews annually to validate user accounts and their associated privileges remain appropriate.",
     "Manual process only. Admin panel provides a user list, but no export, diff, or historical comparison capability.",
     "MEDIUM",
     "Teams provides no automated access review tooling. You must manually screenshot or export the user list, compare against your HR/IdP roster, document findings, and track remediation. This is doable but error-prone and time-consuming.",
     "1) Quarterly (not just annual) access reviews of Claude Teams admin panel.\n2) Export/screenshot user list and compare to IdP active users.\n3) Document review in Drata with reviewer name, date, and findings.\n4) Track and remediate any discrepancies within 5 business days.",
     "Audit logs + Compliance API enable automated access review. Analytics API shows actual usage patterns. SIEM integration provides historical access data.",
     "LOW",
     "Enterprise's Compliance API and audit logs provide the data needed for automated or semi-automated access reviews with full audit trail.",
     "NO"),

    # ── CC7 System Operations & Monitoring ──
    ("section", "CC7 — System Operations & Monitoring"),
    ("CC7.1 / CC4.1.3", "System Monitoring & Alerting",
     "Production systems are monitored and automated alerts are sent to personnel based on pre-configured rules. Events are triaged and escalated per policy.",
     "No audit logs, no alerting, no log export from Claude Teams. You have zero visibility into Claude-specific user activity or security events.",
     "HIGH",
     "Your SOC 2 report requires monitoring of all production systems (CC4.1.3). Claude Teams provides no mechanism to monitor user activity, detect anomalous usage, or generate alerts. You cannot integrate Claude data into your existing monitoring stack (AWS CloudWatch, GuardDuty, etc.).",
     "1) Designate a compliance officer for weekly manual review of Claude Teams admin panel.\n2) Create an internal AI usage log where team members record sensitive interactions.\n3) Establish an AI-specific incident reporting channel.\n4) Document this as a compensating control with clear review cadence.\n5) Reference Anthropic's own monitoring (their SOC 2 covers their infrastructure).",
     "Full audit logs with SIEM integration (Splunk, Datadog, Elastic). Compliance API for programmatic real-time monitoring. Analytics API for usage patterns. Automated alerting on anomalous activity.",
     "LOW",
     "Enterprise provides native monitoring, logging, and SIEM export that directly satisfies CC7.1 and CC4.1.3. No compensating controls needed.",
     "NO"),

    ("CC7.2", "Anomaly Detection",
     "Entity monitors system components for anomalies indicative of malicious acts, natural disasters, and errors.",
     "No visibility into Claude-specific anomalies. Relies entirely on Anthropic's internal monitoring.",
     "HIGH",
     "Without audit logs or SIEM integration, you cannot detect anomalous Claude usage (e.g., unusual query volumes, access from unexpected locations, attempts to extract sensitive data). You are blind to insider threats involving AI tools.",
     "1) SSO conditional access policies to restrict access by IP/location/device.\n2) Manual periodic review of user activity patterns via admin panel.\n3) Employee attestation process for AI usage compliance.\n4) Document reliance on Anthropic's platform-level monitoring in your vendor risk assessment.",
     "Audit logs capture all user actions. SIEM integration enables correlation with other security data sources. IP allowlisting restricts access to approved networks.",
     "LOW",
     "Enterprise audit logs + SIEM integration + IP allowlisting provide comprehensive anomaly detection capability aligned with your existing AWS GuardDuty/Inspector monitoring stack.",
     "NO"),

    ("CC7.3", "Security Event Evaluation",
     "Entity evaluates security events to determine if they could result in failure to meet objectives, and takes action to prevent/address failures.",
     "No Claude-specific security event data available on Teams. Cannot evaluate what you cannot see.",
     "HIGH",
     "This control requires evaluating security events from all systems in scope. If Claude is in scope (and it should be as a tool handling internal workflows), the inability to access event data is a gap. Your existing Incident Response Plan (CC7.3.2) cannot cover Claude events without event data.",
     "1) Add AI-specific scenarios to Incident Response Plan (data leakage via AI, unauthorized access, prompt injection).\n2) Create an AI incident classification framework.\n3) Establish manual reporting procedures for Claude-related security concerns.\n4) Document Anthropic's Trust & Safety review processes as a vendor control.",
     "Compliance API enables evaluation of security events. Audit logs provide event data for incident investigation. Full integration into your incident management workflow.",
     "LOW",
     "Enterprise provides the event data needed to satisfy CC7.3. Incident investigation can use actual audit log data rather than relying on manual reports.",
     "NO"),

    # ── CC3 / CC9 — Risk Assessment & Vendor Management ──
    ("section", "CC3 / CC9 — Risk Assessment & Vendor Management"),
    ("CC3.1.1", "Annual Risk Assessment",
     "Risk assessments conducted annually, including threats, vulnerabilities, likelihood, impact. Risk owners assigned. Results documented in risk register.",
     "Not plan-dependent. This is your internal process. However, adding Claude as a new tool requires updating your risk register.",
     "MEDIUM",
     "The control itself is satisfied by your internal risk assessment process (already passing). The risk is that if you add Claude without updating your risk register, the next audit could flag an incomplete risk assessment. AI-specific risks (data leakage, hallucination, prompt injection) must be assessed.",
     "1) Add Claude/AI as a new risk category in your risk register.\n2) Assess risks: data leakage, unauthorized access, model hallucination, prompt injection, vendor lock-in.\n3) Assign risk owners and treatment options for each.\n4) Include in next annual risk assessment cycle.",
     "Same — this is your internal process regardless of plan.",
     "MEDIUM",
     "Same requirement to update your risk register. Enterprise features reduce the residual risk ratings for each AI risk, but the risk assessment itself must still be performed.",
     "NO"),

    ("CC3.2.2", "Vendor Compliance Review",
     "Compliance reports for critical vendors reviewed at least annually. Results and action items documented.",
     "Anthropic's SOC 2 Type II report available under NDA at trust.anthropic.com. SOC 3 summary publicly available. ISO 27001, ISO 42001, CSA STAR certifications apply to all plans.",
     "LOW",
     "Anthropic's certifications cover the same infrastructure for both Teams and Enterprise. You can obtain their SOC 2 Type II report for your vendor risk assessment. No plan-level gap.",
     "1) Request Anthropic's SOC 2 Type II report annually from trust.anthropic.com.\n2) Document review findings in vendor risk register.\n3) Track any action items from the review.",
     "Same — Anthropic's certifications apply to all plans. Enterprise customers may get access to additional compliance documentation through dedicated support.",
     "LOW",
     "Identical vendor compliance posture. Anthropic's infrastructure certifications are plan-agnostic.",
     "NO"),

    ("CC3.2.3", "Vendor Due Diligence",
     "Due diligence performed prior to engaging with a new vendor. Results including action items documented.",
     "All Anthropic compliance documentation is available: SOC 2/3, ISO 27001, ISO 42001, CSA STAR, DPA, Privacy Policy, Trust Center.",
     "LOW",
     "Sufficient documentation exists to complete due diligence. The DPA is automatically incorporated into Commercial Terms of Service for both plans. Anthropic's Trust Center provides comprehensive security information.",
     "1) Complete vendor due diligence for Anthropic before Claude rollout.\n2) Review DPA, Privacy Policy, Terms of Service.\n3) Document findings and action items in your vendor management system.",
     "Same — plus access to dedicated customer success manager for security questionnaires.",
     "LOW",
     "Identical for due diligence purposes. Enterprise adds a dedicated support channel if you need custom security questionnaires answered.",
     "NO"),

    ("CC9.2.1", "Vendor Lifecycle Management",
     "Documented policy for managing vendor and third-party relationships through entire lifecycle.",
     "DPA automatically incorporated. Privacy Policy publicly available. Terms of Service govern relationship.",
     "LOW",
     "Your existing Vendor Management Policy (CC9.2.1, already passing) covers this. Anthropic provides standard commercial terms. No plan-specific gap.",
     "1) Add Anthropic/Claude to your vendor register with lifecycle tracking.\n2) Schedule annual reviews per your Vendor Management Policy.\n3) Document contract terms, DPA, and renewal dates.",
     "Same — plus ability to negotiate custom terms on sales-assisted Enterprise.",
     "LOW",
     "Identical for standard vendor lifecycle management. Enterprise offers custom contract terms if needed.",
     "NO"),

    # ── CC6.5 / CC6.7 — Data Protection ──
    ("section", "CC6.5 / CC6.7 — Data Protection & Retention"),
    ("CC6.5.1", "Data Disposal",
     "Documented policies and procedures for erasure or destruction of information identified for disposal.",
     "Fixed 30-day backend deletion. Inputs/outputs automatically deleted within 30 days. If AUP review triggered, up to 90 days. Not configurable.",
     "MEDIUM",
     "Teams has a fixed 30-day retention with no ability to shorten it. You cannot implement Zero Data Retention. For a fintech handling Customer Data (PII, transaction data, bank account info per your data classification), this creates a window where data could persist at Anthropic. However, your Acceptable Use Policy should prohibit submitting such data to Claude.",
     "1) Document Anthropic's 30-day retention in your data retention schedule.\n2) Create/update Acceptable Use Policy to prohibit submitting Customer Data and Confidential data to Claude.\n3) Train all employees on data classification before AI use.\n4) Add AI data handling to annual security awareness training (CC1.4.2).\n5) Consider DLP tools to prevent sensitive data from being pasted into Claude.",
     "Configurable retention periods (minimum 30 days). Zero Data Retention (ZDR) available via addendum — no prompts/outputs persisted after response.",
     "LOW",
     "Enterprise gives you full control over retention. ZDR eliminates the data persistence concern entirely. Configurable retention aligns with your existing Data Retention Policy.",
     "NO"),

    ("CC6.7.1", "Data In Transit / At Rest",
     "Protect data during transmission, movement, or removal. OS updates enabled for security patches.",
     "TLS 1.2+ in transit. AES-256 at rest. Confidential inference with TPM root of trust. Available on both plans.",
     "LOW",
     "Anthropic's encryption standards meet or exceed industry requirements. TLS and AES-256 are the same on both plans. This is an infrastructure-level control covered by Anthropic's SOC 2.",
     "N/A — No compensating controls needed. Document Anthropic's encryption standards in your vendor risk assessment.",
     "Same — TLS 1.2+ in transit, AES-256 at rest.",
     "LOW",
     "Identical encryption. Both plans use the same infrastructure.",
     "NO"),

    # ── CC6.6 — Boundary Protection ──
    ("section", "CC6.6 — Boundary Protection"),
    ("CC6.6 (general)", "Network Access Restriction",
     "Logical access security measures to protect against threats from sources outside system boundaries.",
     "SSO with conditional access policies from IdP. No native IP allowlisting or tenant restrictions in Claude Teams.",
     "MEDIUM",
     "Teams lacks IP allowlisting and tenant restrictions. While SSO provides authentication, you cannot restrict Claude access to specific corporate networks or devices natively. Your IdP conditional access policies partially compensate, but Claude itself cannot enforce network-level restrictions.",
     "1) Configure SSO conditional access in your IdP to restrict Claude login by IP range, device compliance, and location.\n2) Enable MFA for all Claude access.\n3) Document IdP-level network controls as compensating for the lack of native IP allowlisting.\n4) Monitor IdP sign-in logs for Claude access from unexpected locations.",
     "IP allowlisting restricts access to approved corporate networks. Tenant restrictions prevent data leakage to personal accounts. SSO + SCIM + audit logs for comprehensive access control.",
     "LOW",
     "Enterprise's native IP allowlisting and tenant restrictions provide defense-in-depth beyond SSO, directly satisfying boundary protection requirements.",
     "NO"),

    # ── CC8 — Change Management ──
    ("section", "CC8 — Change Management"),
    ("CC8.1.1-CC8.1.6", "Change Management (SDLC)",
     "Version control, peer review, segregation of duties, testing in non-prod, approval before production deployment, documented change management policies.",
     "These controls apply to YOUR development process (GitHub), not Claude itself. Claude Code on Teams supports managed policy settings for tool permissions and file access.",
     "LOW",
     "Change management controls in your SOC 2 are about your Verivend application SDLC via GitHub (your subservice org). Claude/Claude Code is a development tool, not the system being changed. As long as your GitHub-based SDLC controls remain intact (peer review, approvals, testing), this is satisfied. Claude Code's managed policy settings allow you to restrict what the tool can do.",
     "1) If using Claude Code for development, ensure all AI-generated code still goes through your existing peer review and approval process.\n2) Configure Claude Code managed policy settings to restrict production access.\n3) Document Claude Code as a development tool in your SDLC policy.",
     "Same — plus enhanced Claude Code analytics (lines accepted, accept rate, usage patterns) for additional visibility into AI-assisted development.",
     "LOW",
     "Both plans support managed policy settings. Enterprise adds analytics for AI code contribution tracking, useful for audit evidence but not strictly required.",
     "NO"),

    # ── CC1 / CC2 — Control Environment & Communication ──
    ("section", "CC1 / CC2 — Control Environment, Communication & Information"),
    ("CC1.4.2", "Security Training",
     "Training programs for information security best practices. Required during onboarding and annually.",
     "Not plan-dependent. This is your internal training program.",
     "LOW",
     "Your existing training program is passing. The gap is ensuring AI-specific content is included. This is a process update, not a tooling issue.",
     "1) Add AI security awareness module to existing onboarding and annual training.\n2) Cover: data classification for AI, acceptable vs prohibited use cases, prompt injection risks, reporting procedures.\n3) Track completion in your training records (Justworks/Drata).",
     "Same — your internal training program.",
     "LOW",
     "Identical requirement regardless of plan. You need AI-specific training content either way.",
     "NO"),

    ("CC2.1.1", "Architectural Diagram",
     "Documented architectural diagram showing system boundaries. Reviewed and approved annually.",
     "Not plan-dependent. Must update your architecture diagram to include Claude as a tool.",
     "LOW",
     "Your existing architectural diagram covers Verivend. If Claude becomes part of your operational toolset, it should be reflected in the diagram showing data flows between users, Claude, and your systems.",
     "1) Update architectural diagram to include Claude as an external service.\n2) Show data flow: User -> SSO/IdP -> Claude -> (no data stored beyond 30 days).\n3) Include in next annual review cycle.",
     "Same — diagram update needed regardless of plan.",
     "LOW",
     "Identical. Architecture documentation is your responsibility on both plans.",
     "NO"),

    ("CC2.1.2", "Information Security Policy",
     "Defined and documented Information Security Policy supporting internal control.",
     "Not plan-dependent. Policy should address AI tool usage.",
     "LOW",
     "Your existing ISP is passing. Adding AI/Claude usage guidelines is a policy update, not a tooling issue.",
     "1) Add AI/Claude section to Information Security Policy.\n2) Address: approved use cases, data restrictions, access controls, monitoring expectations.",
     "Same — policy update needed regardless of plan.",
     "LOW",
     "Identical. Policy updates are your responsibility on both plans.",
     "NO"),
]

r = 5
for row_data in rows1:
    if row_data[0] == "section":
        ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols1)
        ws1.cell(row=r, column=1, value=row_data[1])
        style_row(ws1, r, ncols1, is_section=True)
    else:
        ref, control, requirement, teams_cap, teams_risk, risk_rationale, comp_controls, ent_cap, ent_risk, ent_rationale, exception = row_data
        vals = [ref, control, requirement, teams_cap, teams_risk, risk_rationale, comp_controls, ent_cap, ent_risk, ent_rationale, exception]
        for c, v in enumerate(vals, 1):
            ws1.cell(row=r, column=c, value=v)
        style_row(ws1, r, ncols1)
        apply_risk_format(ws1.cell(row=r, column=5), teams_risk)
        apply_risk_format(ws1.cell(row=r, column=9), ent_risk)
        # Highlight exceptions
        exc_cell = ws1.cell(row=r, column=11)
        if exception == "YES":
            exc_cell.fill = high_fill
            exc_cell.font = high_font
    r += 1

ws1.auto_filter.ref = f"A4:K{r-1}"
ws1.freeze_panes = "A5"


# ════════════════════════════════════════════════════════════════════
# TAB 2 — Risk Summary
# ════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Risk Summary")
ws2.sheet_properties.tabColor = "C00000"

ws2.merge_cells("A1:F1")
ws2["A1"].value = "Risk Summary — Teams vs Enterprise"
ws2["A1"].font = title_font

headers2 = ["Risk Category", "# Controls", "Teams HIGH", "Teams MEDIUM", "Teams LOW", "Enterprise HIGH", "Enterprise MEDIUM", "Enterprise LOW"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
style_header(ws2, 3, len(headers2))

risk_summary = [
    ("Logical & Physical Access (CC5/CC6)", 8, 2, 2, 4, 0, 0, 8),
    ("System Operations & Monitoring (CC7)", 3, 3, 0, 0, 0, 0, 3),
    ("Risk Assessment & Vendor Mgmt (CC3/CC9)", 4, 0, 1, 3, 0, 1, 3),
    ("Data Protection & Retention (CC6.5/CC6.7)", 2, 0, 1, 1, 0, 0, 2),
    ("Boundary Protection (CC6.6)", 1, 0, 1, 0, 0, 0, 1),
    ("Change Management (CC8)", 1, 0, 0, 1, 0, 0, 1),
    ("Control Environment (CC1/CC2)", 3, 0, 0, 3, 0, 0, 3),
    ("TOTALS", 22, 5, 5, 12, 0, 1, 21),
]

for ri, row in enumerate(risk_summary, 4):
    for ci, val in enumerate(row, 1):
        cell = ws2.cell(row=ri, column=ci, value=val)
        cell.border = thin_border
        cell.alignment = wrap
        if ri == len(risk_summary) + 3:
            cell.font = bold_font
        else:
            cell.font = normal_font
    # Color the HIGH counts
    if row[2] and row[2] > 0:
        c = ws2.cell(row=ri, column=3)
        c.fill = high_fill
        c.font = high_font
    if row[5] and row[5] > 0:
        c = ws2.cell(row=ri, column=6)
        c.fill = high_fill
        c.font = high_font

for i, w in enumerate([40, 14, 14, 16, 14, 16, 18, 16], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Add key insight
r_insight = len(risk_summary) + 5
ws2.merge_cells(f"A{r_insight}:H{r_insight}")
ws2.cell(row=r_insight, column=1, value="KEY INSIGHT: Teams plan results in 5 HIGH-risk controls and 5 MEDIUM-risk controls. Enterprise eliminates ALL high-risk controls, reducing to 0 HIGH and 1 MEDIUM (risk assessment update, which is required on both plans).")
ws2.cell(row=r_insight, column=1).font = Font(name="Calibri", bold=True, size=11, color="C00000")
ws2.cell(row=r_insight, column=1).alignment = wrap


# ════════════════════════════════════════════════════════════════════
# TAB 3 — Feature Comparison
# ════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Feature Comparison")
ws3.sheet_properties.tabColor = "548235"

ws3.merge_cells("A1:E1")
ws3["A1"].value = "Claude Teams vs Enterprise — Full Feature Comparison"
ws3["A1"].font = title_font

headers3 = ["Feature", "Teams", "Enterprise", "SOC 2 Relevance", "Impact on iAltA Compliance"]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=i, value=h)
style_header(ws3, 3, len(headers3))

features = [
    ("SSO (SAML 2.0 / OIDC)", "Yes", "Yes", "CC6.1.3, CC6.6", "No gap — both plans support SSO"),
    ("Domain Capture", "Yes", "Yes", "CC6.1.6", "No gap — prevents shadow IT"),
    ("JIT Provisioning", "Yes", "Yes", "CC5.1.1", "Partial — automates onboarding but doesn't document approval"),
    ("SCIM 2.0 Provisioning", "No", "Yes", "CC5.1.1, CC6.1.2, CC6.2", "CRITICAL GAP on Teams — directly related to both audit exceptions"),
    ("Role-Based Permissioning", "Basic (4 roles)", "Fine-grained + SCIM-synced", "CC6.1.7, CC6.3", "Medium gap — Teams RBAC is functional but limited"),
    ("Audit Logs", "No", "Yes (SIEM export: Splunk, Datadog, Elastic)", "CC7.1, CC7.2, CC7.3, CC4.1.3", "CRITICAL GAP on Teams — no visibility into user activity"),
    ("Compliance API", "No", "Yes", "CC7.1, CC4.1", "CRITICAL GAP — no programmatic compliance monitoring"),
    ("Analytics API", "No", "Yes", "CC4.1, CC7.2", "Gap on Teams — no usage analytics for audit evidence"),
    ("Custom Data Retention", "No (fixed 30 days)", "Yes (configurable, min 30 days)", "CC6.5.1", "Medium gap — cannot customize retention to match your policy"),
    ("Zero Data Retention (ZDR)", "No", "Yes (via addendum)", "CC6.5.1", "Gap — Teams data persists for 30 days minimum"),
    ("IP Allowlisting", "No", "Yes", "CC6.6", "Medium gap — must use IdP conditional access as compensating control"),
    ("Tenant Restrictions", "No", "Yes", "CC6.6, CC6.7", "Gap — cannot prevent data leakage to personal Claude accounts"),
    ("HIPAA / BAA", "No", "Sales-assisted only", "N/A (not in current scope)", "Not applicable to current SOC 2 scope (Security only)"),
    ("Zero Training on Data", "Yes", "Yes", "CC6.5, CC6.7", "No gap — Anthropic does not train on commercial data"),
    ("Encryption (TLS 1.2+ / AES-256)", "Yes", "Yes", "CC6.7", "No gap — same infrastructure encryption"),
    ("Spend Controls", "Per-seat limits", "Org-level + per-user caps", "N/A", "Not SOC 2 relevant, but useful for governance"),
    ("Claude Code", "Premium seats only ($100-125/seat/mo)", "All seats (included)", "CC8.1", "Cost consideration — Enterprise includes Claude Code at no extra per-seat cost"),
    ("Cowork", "No", "All seats (included)", "N/A", "Not SOC 2 relevant"),
    ("Connectors (Google, GitHub, Slack, M365)", "Yes", "Yes", "CC6.7, CC9.2", "No gap — same connector availability"),
    ("Dedicated Support", "No", "Sales-assisted", "N/A", "Enterprise provides a point of contact for compliance questions"),
    ("Max Seats", "150", "Unlimited (min 20)", "N/A", "Not currently a constraint at 38 users"),
    ("Context Window", "200K tokens", "500K-1M tokens", "N/A", "Not SOC 2 relevant, but affects Claude Code capability"),
]

for ri, row in enumerate(features, 4):
    for ci, val in enumerate(row, 1):
        cell = ws3.cell(row=ri, column=ci, value=val)
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = normal_font
        if "CRITICAL GAP" in str(val):
            cell.fill = high_fill
            cell.font = high_font

for i, w in enumerate([30, 25, 35, 25, 45], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.auto_filter.ref = f"A3:E{3 + len(features)}"
ws3.freeze_panes = "A4"


# ════════════════════════════════════════════════════════════════════
# TAB 4 — Pricing Comparison
# ════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Pricing Breakdown")
ws4.sheet_properties.tabColor = "BF8F00"

ws4.merge_cells("A1:F1")
ws4["A1"].value = "Pricing Comparison — Teams vs Enterprise (38 Users)"
ws4["A1"].font = title_font

# Teams pricing
ws4.cell(row=3, column=1, value="TEAMS PLAN PRICING").font = subtitle_font
ws4.merge_cells("A3:F3")

teams_headers = ["Configuration", "Seat Cost/Mo", "# Users", "Monthly Cost", "Annual Cost", "Notes"]
for i, h in enumerate(teams_headers, 1):
    ws4.cell(row=4, column=i, value=h)
style_header(ws4, 4, 6)

teams_rows = [
    ("Standard Only (monthly billing)", "$25/seat", 38, "$950", "$11,400", "No Claude Code access"),
    ("Standard Only (annual billing)", "$20/seat", 38, "$760", "$9,120", "No Claude Code access"),
    ("All Premium (monthly billing)", "$125/seat", 38, "$4,750", "$57,000", "All users get Claude Code"),
    ("All Premium (annual billing)", "$100/seat", 38, "$3,800", "$45,600", "All users get Claude Code"),
    ("Mixed: 30 Standard + 8 Premium (annual)", "Mixed", 38, "$1,400", "$16,800", "8 devs with Claude Code @ $100 + 30 standard @ $20"),
    ("Mixed: 30 Standard + 8 Premium (monthly)", "Mixed", 38, "$1,750", "$21,000", "8 devs with Claude Code @ $125 + 30 standard @ $25"),
]

for ri, row in enumerate(teams_rows, 5):
    for ci, val in enumerate(row, 1):
        cell = ws4.cell(row=ri, column=ci, value=val)
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = normal_font

# Enterprise pricing
r = 5 + len(teams_rows) + 1
ws4.cell(row=r, column=1, value="ENTERPRISE PLAN PRICING (Self-Serve)").font = subtitle_font
ws4.merge_cells(f"A{r}:F{r}")
r += 1

ent_headers = ["Component", "Cost", "# Users", "Monthly Cost", "Annual Cost", "Notes"]
for i, h in enumerate(ent_headers, 1):
    ws4.cell(row=r, column=i, value=h)
style_header(ws4, r, 6)
r += 1

ent_rows = [
    ("Seat Cost (annual billing)", "$20/seat/mo", 38, "$760", "$9,120", "Includes Claude web/desktop/mobile, Claude Code, and Cowork for ALL seats"),
    ("Usage (token consumption)", "API rates", "N/A", "Variable", "Variable", "Billed at standard API rates. Depends on actual usage."),
    ("Per-User Spend Limit", "Configurable", "N/A", "N/A", "N/A", "Options: $20, $50, $100, $200, $500, or $1,000/mo per user"),
    ("Estimated Total (moderate usage)", "", 38, "~$1,500-2,500", "~$18,000-30,000", "Seat cost + estimated token usage for 38 users with moderate daily use"),
    ("Estimated Total (heavy usage)", "", 38, "~$3,000-5,000", "~$36,000-60,000", "Seat cost + estimated token usage for heavy Claude Code + chat usage"),
]

for row in ent_rows:
    for ci, val in enumerate(row, 1):
        cell = ws4.cell(row=r, column=ci, value=val)
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = normal_font
    r += 1

for i, w in enumerate([40, 18, 12, 16, 16, 50], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w


# ════════════════════════════════════════════════════════════════════
# TAB 5 — Cost-Benefit Analysis
# ════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Cost-Benefit Analysis")
ws5.sheet_properties.tabColor = "7030A0"

ws5.merge_cells("A1:E1")
ws5["A1"].value = "Cost-Benefit Analysis — Teams vs Enterprise for SOC 2 Compliance"
ws5["A1"].font = title_font

# Direct cost comparison
ws5.cell(row=3, column=1, value="DIRECT COST COMPARISON (38 Users, Annual)").font = subtitle_font
ws5.merge_cells("A3:E3")

cost_headers = ["Cost Category", "Teams (Standard Annual)", "Teams (Mixed: 30 Std + 8 Premium Annual)", "Enterprise (Self-Serve, Moderate Usage)", "Enterprise (Self-Serve, Heavy Usage)"]
for i, h in enumerate(cost_headers, 1):
    ws5.cell(row=4, column=i, value=h)
style_header(ws5, 4, 5)

cost_rows = [
    ("Seat Costs (Annual)", "$9,120", "$16,800", "$9,120", "$9,120"),
    ("Usage/Token Costs", "Included in seat", "Included in seat", "~$9,000-$21,000", "~$27,000-$51,000"),
    ("Claude Code Access", "Not included", "$9,600 (8 Premium upgrades)", "Included for all 38 users", "Included for all 38 users"),
    ("Cowork Access", "Not available", "Not available", "Included for all 38 users", "Included for all 38 users"),
    ("Subtotal: Platform Cost", "$9,120/yr", "$16,800/yr", "~$18,000-$30,000/yr", "~$36,000-$60,000/yr"),
]

r = 5
for row in cost_rows:
    for ci, val in enumerate(row, 1):
        cell = ws5.cell(row=r, column=ci, value=val)
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = bold_font if "Subtotal" in row[0] else normal_font
    r += 1

# Compliance overhead
r += 1
ws5.cell(row=r, column=1, value="COMPLIANCE OVERHEAD (Annual Estimate)").font = subtitle_font
ws5.merge_cells(f"A{r}:E{r}")
r += 1

comp_headers = ["Compliance Activity", "Teams (Hours/Year)", "Teams (Est. Cost @ $75/hr)", "Enterprise (Hours/Year)", "Enterprise (Est. Cost @ $75/hr)"]
for i, h in enumerate(comp_headers, 1):
    ws5.cell(row=r, column=i, value=h)
style_header(ws5, r, 5)
r += 1

comp_rows = [
    ("Weekly admin panel reviews (CC7.1/CC7.2)", "104 hrs (2 hrs/week)", "$7,800", "0 hrs (automated)", "$0"),
    ("Monthly access reconciliation (CC5.1.1/CC6.1.2)", "24 hrs (2 hrs/month)", "$1,800", "0 hrs (SCIM automated)", "$0"),
    ("Quarterly access reviews (CC4.1.2)", "16 hrs (4 hrs/quarter)", "$1,200", "4 hrs (Compliance API assisted)", "$300"),
    ("Onboarding/offboarding documentation (CC5.1.1/CC6.1.2)", "38 hrs (1 hr per user event)", "$2,850", "2 hrs (SCIM automated, spot checks)", "$150"),
    ("AI usage tracker maintenance", "52 hrs (1 hr/week)", "$3,900", "0 hrs (audit logs)", "$0"),
    ("Quarterly employee attestations", "16 hrs (4 hrs/quarter)", "$1,200", "0 hrs (not needed with audit logs)", "$0"),
    ("Compensating control documentation for auditor", "40 hrs/year", "$3,000", "0 hrs (native controls)", "$0"),
    ("Incident response for AI events (manual investigation)", "20 hrs/year (est.)", "$1,500", "4 hrs/year (log-based investigation)", "$300"),
    ("Annual training content update (AI module)", "8 hrs", "$600", "8 hrs", "$600"),
    ("Vendor risk assessment (Anthropic)", "4 hrs", "$300", "4 hrs", "$300"),
    ("SUBTOTAL: Compliance Labor", "322 hrs/year", "$24,150/yr", "22 hrs/year", "$1,650/yr"),
]

for row in comp_rows:
    for ci, val in enumerate(row, 1):
        cell = ws5.cell(row=r, column=ci, value=val)
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = bold_font if "SUBTOTAL" in row[0] else normal_font
    r += 1

# Total cost of ownership
r += 1
ws5.cell(row=r, column=1, value="TOTAL COST OF OWNERSHIP (Annual)").font = subtitle_font
ws5.merge_cells(f"A{r}:E{r}")
r += 1

tco_headers = ["Component", "Teams (Std Annual)", "Teams (Mixed Annual)", "Enterprise (Moderate)", "Enterprise (Heavy)"]
for i, h in enumerate(tco_headers, 1):
    ws5.cell(row=r, column=i, value=h)
style_header(ws5, r, 5)
r += 1

tco_rows = [
    ("Platform Cost", "$9,120", "$16,800", "~$18,000-$30,000", "~$36,000-$60,000"),
    ("Compliance Labor", "$24,150", "$24,150", "$1,650", "$1,650"),
    ("Audit Risk Premium (potential re-audit, exception remediation)", "$5,000-$15,000", "$5,000-$15,000", "$0", "$0"),
    ("TOTAL ESTIMATED TCO", "$38,270-$48,270", "$45,950-$55,950", "~$19,650-$31,650", "~$37,650-$61,650"),
]

for row in tco_rows:
    for ci, val in enumerate(row, 1):
        cell = ws5.cell(row=r, column=ci, value=val)
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = bold_font if "TOTAL" in row[0] else normal_font
        if "TOTAL" in row[0]:
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    r += 1

# Key takeaways
r += 2
ws5.cell(row=r, column=1, value="KEY TAKEAWAYS").font = subtitle_font
r += 1
takeaways = [
    "1. Teams Standard is the cheapest platform cost ($9,120/yr) but has the highest compliance overhead ($24,150/yr in labor + $5K-$15K audit risk).",
    "2. Enterprise Moderate Usage (~$18K-$30K/yr platform) has the lowest total cost of ownership when compliance labor and audit risk are factored in (~$19,650-$31,650 total).",
    "3. The compliance labor savings on Enterprise (~$22,500/yr) can partially or fully offset the higher token consumption costs.",
    "4. Enterprise eliminates the audit risk premium entirely — no compensating controls means no risk of auditor rejecting your approach.",
    "5. Enterprise includes Claude Code for all seats (no Premium upgrade needed) and Cowork — a significant value add vs Teams Mixed pricing.",
    "6. CRITICAL: Given your qualified opinion with two access control exceptions, the audit risk premium on Teams could be significantly higher if exceptions recur.",
    "7. Migration from Teams to Enterprise is one-way and irreversible. Factor this into your decision timeline.",
]

for t in takeaways:
    ws5.cell(row=r, column=1, value=t).font = normal_font
    ws5.cell(row=r, column=1).alignment = wrap
    ws5.merge_cells(f"A{r}:E{r}")
    r += 1

for i, w in enumerate([45, 22, 28, 28, 28], 1):
    ws5.column_dimensions[get_column_letter(i)].width = w


# ════════════════════════════════════════════════════════════════════
# TAB 6 — Compensating Controls Detail
# ════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Compensating Controls")
ws6.sheet_properties.tabColor = "ED7D31"

ws6.merge_cells("A1:G1")
ws6["A1"].value = "Compensating Controls Required on Teams Plan"
ws6["A1"].font = title_font

ws6.merge_cells("A2:G2")
ws6["A2"].value = "These controls are ONLY required if staying on the Teams plan. Enterprise plan satisfies these requirements natively."
ws6["A2"].font = Font(name="Calibri", italic=True, size=10, color="C00000")

headers6 = ["SOC 2 Ref", "Gap on Teams", "Compensating Control", "Responsible Party", "Frequency", "Evidence Required", "Priority"]
for i, h in enumerate(headers6, 1):
    ws6.cell(row=4, column=i, value=h)
style_header(ws6, 4, 7)

cc_rows = [
    ("CC5.1.1", "No SCIM — no automated access provisioning documentation",
     "Create documented onboarding checklist. Each Claude access grant requires a Jira/Drata ticket with manager approval before provisioning.",
     "IT Admin / People Ops", "Per onboarding event",
     "Jira ticket with manager approval, screenshot of admin panel showing user added, date stamp",
     "CRITICAL"),
    ("CC5.1.1", "No audit log of access approvals",
     "Export Claude Teams admin panel user list monthly. Cross-reference with approved access request tickets. Store in Drata.",
     "Compliance Officer", "Monthly",
     "Monthly reconciliation report with reviewer sign-off, list of discrepancies and remediation actions",
     "CRITICAL"),
    ("CC6.1.2", "No SCIM — no automated deprovisioning",
     "Add Claude to offboarding checklist with <1 business day SLA. Leverage IdP SSO session invalidation. Screenshot admin panel after removal.",
     "IT Admin / People Ops", "Per termination event",
     "Offboarding ticket with timestamp, screenshot of user removal from admin panel, IdP deactivation record",
     "CRITICAL"),
    ("CC6.1.2", "No audit trail proving deprovisioning timing",
     "Weekly terminated employee list vs active Claude users sweep. Document any gaps and remediation.",
     "IT Admin", "Weekly",
     "Weekly sweep log with date, reviewer name, findings, and remediation actions if any",
     "CRITICAL"),
    ("CC7.1/CC4.1.3", "No audit logs or monitoring",
     "Designate compliance officer to review Claude Teams admin panel weekly. Document review findings.",
     "Compliance Officer", "Weekly",
     "Weekly review log with date, reviewer name, observations, and any flagged activity",
     "HIGH"),
    ("CC7.2", "No anomaly detection for Claude activity",
     "Configure SSO conditional access policies (IP/location/device restrictions). Monitor IdP sign-in logs for Claude access anomalies.",
     "Security Engineer", "Ongoing + Monthly review",
     "IdP conditional access policy config, monthly review of sign-in logs, flagged anomalies",
     "HIGH"),
    ("CC7.3", "No security event data from Claude",
     "Add AI-specific scenarios to Incident Response Plan. Create manual incident reporting channel for Claude-related concerns.",
     "Security / Compliance", "Annual update + per-incident",
     "Updated IRP document, incident reports, AI incident classification records",
     "HIGH"),
    ("CC6.5.1", "No custom data retention (fixed 30 days)",
     "Document Anthropic's 30-day retention in data retention schedule. Create Acceptable Use Policy prohibiting Customer Data/Confidential data in Claude.",
     "Compliance Officer", "Annual policy review",
     "Updated data retention schedule, signed AUP acknowledgments from all employees",
     "MEDIUM"),
    ("CC6.5.1", "No Zero Data Retention option",
     "Train all employees on data classification. Prohibit PII, financial data, transaction data in Claude prompts.",
     "All employees / Compliance", "Annual training + ongoing",
     "Training completion records, AUP acknowledgments, data classification guide",
     "MEDIUM"),
    ("CC6.6", "No IP allowlisting",
     "Configure IdP conditional access to restrict Claude login by IP range, device compliance, and location. Enable MFA.",
     "Security Engineer", "Initial config + quarterly review",
     "IdP conditional access policy configuration, quarterly review of policy effectiveness",
     "MEDIUM"),
    ("CC6.1.7", "Basic RBAC only (4 roles)",
     "Document role assignment rationale for each Admin/Owner. Minimize privileged role count. Include in annual access review.",
     "IT Admin / Compliance", "Annual + per role change",
     "Role assignment matrix with justification, annual review documentation",
     "MEDIUM"),
    ("CC4.1.2", "No automated access review tooling",
     "Quarterly access reviews: export user list, compare to IdP, document findings, remediate within 5 business days.",
     "Compliance Officer", "Quarterly",
     "Quarterly review report with user list comparison, findings, remediation actions",
     "MEDIUM"),
    ("CC8.1", "No Claude Code analytics",
     "Document that all AI-generated code goes through existing GitHub peer review and approval process. Configure Claude Code managed policy settings.",
     "Engineering Lead", "Ongoing + annual policy review",
     "Claude Code policy config, sample of PRs showing peer review of AI-generated code",
     "LOW"),
]

for ri, row in enumerate(cc_rows, 5):
    for ci, val in enumerate(row, 1):
        cell = ws6.cell(row=ri, column=ci, value=val)
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = normal_font
    # Format priority
    p_cell = ws6.cell(row=ri, column=7)
    if row[6] == "CRITICAL":
        p_cell.fill = high_fill
        p_cell.font = high_font
    elif row[6] == "HIGH":
        p_cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        p_cell.font = Font(name="Calibri", bold=True, color="C00000")
    elif row[6] == "MEDIUM":
        p_cell.fill = med_fill
        p_cell.font = med_font
    elif row[6] == "LOW":
        p_cell.fill = low_fill
        p_cell.font = low_font

for i, w in enumerate([14, 35, 50, 22, 22, 45, 14], 1):
    ws6.column_dimensions[get_column_letter(i)].width = w

ws6.auto_filter.ref = f"A4:G{4 + len(cc_rows)}"
ws6.freeze_panes = "A5"


# ════════════════════════════════════════════════════════════════════
# TAB 7 — Audit Exceptions
# ════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Audit Exceptions")
ws7.sheet_properties.tabColor = "FF0000"

ws7.merge_cells("A1:G1")
ws7["A1"].value = "SOC 2 Type 2 Audit Exceptions — Remediation & Claude Impact Analysis"
ws7["A1"].font = title_font

ws7.merge_cells("A2:G2")
ws7["A2"].value = "Qualified Opinion Basis — Richey May, February 23, 2026"
ws7["A2"].font = Font(name="Calibri", italic=True, size=10, color="C00000")

headers7 = ["Ref", "Control", "Exception Detail", "Management Response", "Claude Teams Impact", "Claude Enterprise Impact", "Remediation Status"]
for i, h in enumerate(headers7, 1):
    ws7.cell(row=4, column=i, value=h)
style_header(ws7, 4, 7)

exc_rows = [
    ("CC5.1.1",
     "Access requests to information resources are documented and approved by Management based on least privilege, need to know, and segregation of duties principles.",
     "Management was unable to provide supporting documentation for the access request and approval for the samples selected, as the formalized process for documenting and approving access requests was not implemented until December 2025.",
     "An updated process and tooling was put in place to track requests, and employees were trained on proper protocols.",
     "HIGH RISK: Teams lacks SCIM and audit logs. Adding Claude as another system without automated provisioning documentation compounds this exception. The auditor will specifically test whether the 'updated process and tooling' generates proper evidence for ALL systems, including Claude. Manual processes for Claude access documentation are fragile and audit-prone.",
     "LOW RISK: Enterprise SCIM auto-documents provisioning events. Audit logs capture all access changes with timestamps. This directly generates the 'supporting documentation' the auditor found missing.",
     "In progress — process implemented December 2025. Next audit will test effectiveness."),
    ("CC6.1.2",
     "System and physical access is revoked within one business day of effective termination date for terminated users (including employees, third parties and vendors, and other personnel).",
     "Per inspection, it was noted physical access was revoked within one business day of termination; however, system access was not revoked within one business day of termination. Additionally, supporting documentation for the access removal ticket was unable to be obtained.",
     "An updated process and tooling was put in place to track terminations, and people managers were trained on proper protocols.",
     "HIGH RISK: Teams lacks SCIM for automated deprovisioning. Without SCIM, revoking Claude access depends on someone remembering to manually remove the user within 1 business day AND documenting it. This is the exact failure mode that caused the original exception. Adding another manual deprovisioning step increases the risk of recurrence.",
     "LOW RISK: Enterprise SCIM auto-deprovisions when user is removed from IdP group. The access removal is immediate and automatically logged. This eliminates the manual step that caused the exception.",
     "In progress — process implemented post-audit period. Next audit will test effectiveness."),
]

for ri, row in enumerate(exc_rows, 5):
    for ci, val in enumerate(row, 1):
        cell = ws7.cell(row=ri, column=ci, value=val)
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = normal_font
    # Highlight impact cells
    teams_cell = ws7.cell(row=ri, column=5)
    teams_cell.fill = high_fill
    teams_cell.font = high_font
    ent_cell = ws7.cell(row=ri, column=6)
    ent_cell.fill = low_fill
    ent_cell.font = low_font

for i, w in enumerate([12, 35, 40, 30, 45, 45, 25], 1):
    ws7.column_dimensions[get_column_letter(i)].width = w


# ════════════════════════════════════════════════════════════════════
# TAB 8 — Action Items
# ════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Action Items")
ws8.sheet_properties.tabColor = "0070C0"

ws8.merge_cells("A1:H1")
ws8["A1"].value = "Action Items — SOC 2 Compliance for Claude Adoption"
ws8["A1"].font = title_font

# Immediate actions (regardless of plan)
ws8.merge_cells("A3:H3")
ws8.cell(row=3, column=1, value="IMMEDIATE ACTIONS (Required Regardless of Plan Choice)").font = subtitle_font

headers8 = ["#", "Action Item", "Description", "SOC 2 Ref", "Owner", "Deadline", "Status", "Notes"]
for i, h in enumerate(headers8, 1):
    ws8.cell(row=4, column=i, value=h)
style_header(ws8, 4, 8)

immediate = [
    (1, "Remediate existing audit exceptions",
     "Verify that the 'updated process and tooling' for access provisioning (CC5.1.1) and deprovisioning (CC6.1.2) are fully operational and generating evidence. These are the basis of your qualified opinion.",
     "CC5.1.1, CC6.1.2", "IT Admin / Compliance", "Immediate", "In Progress",
     "Management responses in Section V indicate remediation started. Must be generating evidence BEFORE next audit."),
    (2, "Complete vendor due diligence for Anthropic",
     "Perform formal vendor due diligence before Claude rollout. Review DPA, Privacy Policy, ToS, Trust Center. Document findings in your vendor management system.",
     "CC3.2.3", "Compliance Officer", "Before Claude rollout", "Not Started",
     "Required by your existing Vendor Management Policy (CC9.2.1)."),
    (3, "Obtain Anthropic's SOC 2 Type II report",
     "Request Anthropic's SOC 2 Type II report under NDA from trust.anthropic.com. Review and document findings in vendor risk register.",
     "CC3.2.2", "Compliance Officer", "Before Claude rollout", "Not Started",
     "SOC 3 summary is publicly available. Full Type II requires NDA."),
    (4, "Add Anthropic/Claude to vendor risk register",
     "Register Anthropic as a subprocessor. Document service description, data categories processed, compliance certifications, contract terms, and risk rating.",
     "CC9.2.1", "Compliance Officer", "Before Claude rollout", "Not Started",
     "Include in Drata or equivalent GRC platform."),
    (5, "Update risk assessment to include AI risks",
     "Add AI tool risks to your annual risk assessment: data leakage, unauthorized access, hallucination, prompt injection, vendor dependency. Assign risk owners and treatment options.",
     "CC3.1.1", "Risk Owner (assigned)", "Next risk assessment cycle", "Not Started",
     "Your risk assessment is annual. If the next one is before Claude rollout, include AI risks."),
    (6, "Create AI Acceptable Use Policy",
     "Define approved and prohibited use cases for Claude. Specify data classification rules: what data may and may not be submitted. Require employee acknowledgment.",
     "CC6.5.1, CC5.3.1", "Compliance / Legal", "Before Claude rollout", "Not Started",
     "Can be an addendum to existing Acceptable Use Policy (CC5.3.1)."),
    (7, "Update Information Security Policy",
     "Add AI/Claude section addressing: approved tools, access controls, data restrictions, monitoring expectations, incident reporting.",
     "CC2.1.2", "Compliance Officer", "Before Claude rollout", "Not Started",
     "Policy already exists and is passing. This is an update, not new creation."),
    (8, "Add AI module to security training",
     "Update onboarding and annual training to include AI security awareness: data classification for AI, acceptable use, prompt injection risks, reporting procedures.",
     "CC1.4.2", "HR / Compliance", "Before next training cycle", "Not Started",
     "Track completion in Justworks/Drata as with existing training."),
    (9, "Update architectural diagram",
     "Add Claude as an external service in your system architectural diagram. Show data flows: User -> SSO/IdP -> Claude. Include data retention notation (30 days or ZDR).",
     "CC2.1.1", "Engineering", "Next annual review", "Not Started",
     "Diagram is reviewed annually per CC2.1.1."),
    (10, "Brief Board of Directors",
     "Present AI adoption plan, compliance implications, and plan recommendation (Teams vs Enterprise) at next quarterly Board meeting.",
     "CC1.2.2", "Management / CISO", "Next quarterly meeting", "Not Started",
     "Board meets quarterly per CC1.2.2. Include risk assessment results."),
]

r = 5
for row in immediate:
    for ci, val in enumerate(row, 1):
        cell = ws8.cell(row=r, column=ci, value=val)
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = normal_font
    r += 1

# Plan decision action
r += 1
ws8.merge_cells(f"A{r}:H{r}")
ws8.cell(row=r, column=1, value="PLAN DECISION").font = subtitle_font
r += 1

for i, h in enumerate(headers8, 1):
    ws8.cell(row=r, column=i, value=h)
style_header(ws8, r, 8)
r += 1

decision_rows = [
    (11, "Make Teams vs Enterprise decision",
     "Review this workbook with stakeholders. Decide between Teams (with compensating controls) or Enterprise (native compliance). Factor in TCO analysis, audit risk, and operational overhead.",
     "All", "Management / CISO", "Before Claude rollout", "Not Started",
     "Migration from Teams to Enterprise is one-way. If starting on Teams, you can upgrade later but not downgrade."),
    (12, "If Teams: Implement all compensating controls",
     "Implement every control listed in the 'Compensating Controls' tab. Document all controls in Drata. Ensure evidence generation is operational BEFORE the next audit period.",
     "CC5.1.1, CC6.1.2, CC7.1-CC7.3, CC6.5, CC6.6", "IT / Compliance", "Before next audit period", "N/A",
     "See 'Compensating Controls' tab for full detail. Estimated 322 hrs/year of ongoing compliance labor."),
    (13, "If Enterprise: Purchase and configure",
     "Purchase Enterprise self-serve (20 seat minimum, $20/seat/mo annual). Configure SCIM with IdP, enable audit logs, set up SIEM integration, configure IP allowlisting, set retention policy.",
     "CC5.1.1, CC6.1.2, CC7.1-CC7.3", "IT Admin", "Before Claude rollout", "N/A",
     "Enterprise includes all compliance features natively. Configuration is one-time."),
    (14, "Configure Claude Code managed policies",
     "Set up managed policy settings for Claude Code: restrict production access, define tool permissions, configure file access restrictions, set MCP server configs.",
     "CC8.1, CC5.1.2", "Engineering Lead", "Before developer rollout", "N/A",
     "Available on both Teams and Enterprise plans."),
]

for row in decision_rows:
    for ci, val in enumerate(row, 1):
        cell = ws8.cell(row=r, column=ci, value=val)
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = normal_font
    r += 1

# Ongoing actions
r += 1
ws8.merge_cells(f"A{r}:H{r}")
ws8.cell(row=r, column=1, value="ONGOING ACTIONS (Post-Rollout)").font = subtitle_font
r += 1

for i, h in enumerate(headers8, 1):
    ws8.cell(row=r, column=i, value=h)
style_header(ws8, r, 8)
r += 1

ongoing_rows = [
    (15, "Annual vendor compliance review",
     "Review Anthropic's SOC 2 Type II report annually. Update vendor risk register. Document review findings and action items.",
     "CC3.2.2", "Compliance Officer", "Annually", "Recurring",
     "Schedule to align with your existing vendor review cycle."),
    (16, "Annual/quarterly access reviews",
     "Review Claude user accounts and role assignments. Verify against IdP active user list. Document and remediate discrepancies.",
     "CC4.1.2", "IT Admin / Compliance", "Quarterly (Teams) or Annually (Enterprise)", "Recurring",
     "Enterprise audit logs make this largely automated."),
    (17, "Annual risk assessment update",
     "Update AI risk categories in annual risk assessment. Reassess likelihood/impact based on actual experience. Update treatment options.",
     "CC3.1.1", "Risk Owner", "Annually", "Recurring",
     "Include any AI-related incidents or near-misses from the year."),
    (18, "Annual training refresh",
     "Update AI security training content. Ensure all employees complete refresher. Track completion.",
     "CC1.4.2", "HR / Compliance", "Annually", "Recurring",
     "Can be combined with existing annual security training."),
    (19, "Ongoing monitoring (Teams only)",
     "Weekly admin panel reviews, monthly access reconciliations, quarterly attestations. ONLY required on Teams plan.",
     "CC7.1, CC7.2", "Compliance Officer", "Weekly/Monthly/Quarterly", "Recurring (Teams only)",
     "Not needed on Enterprise — audit logs and SIEM handle this."),
    (20, "Incident response testing",
     "Include AI-specific scenarios in annual Incident Response Plan test (CC7.3.1). Update plan based on results.",
     "CC7.3.1", "Security / Compliance", "Annually", "Recurring",
     "You already test your IRP annually. Add AI scenarios to the test."),
]

for row in ongoing_rows:
    for ci, val in enumerate(row, 1):
        cell = ws8.cell(row=r, column=ci, value=val)
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = normal_font
    r += 1

for i, w in enumerate([5, 30, 50, 20, 20, 22, 14, 45], 1):
    ws8.column_dimensions[get_column_letter(i)].width = w

ws8.auto_filter.ref = f"A4:H{r-1}"
ws8.freeze_panes = "A5"


# ════════════════════════════════════════════════════════════════════
# TAB 9 — Anthropic Certifications
# ════════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("Anthropic Certifications")
ws9.sheet_properties.tabColor = "548235"

ws9.merge_cells("A1:E1")
ws9["A1"].value = "Anthropic Compliance Certifications & Resources"
ws9["A1"].font = title_font

ws9.merge_cells("A2:E2")
ws9["A2"].value = "These certifications apply to ALL commercial plans (Teams and Enterprise) — same infrastructure"
ws9["A2"].font = Font(name="Calibri", italic=True, size=10, color="006100")

headers9 = ["Certification", "Status", "Applies To", "How to Obtain", "SOC 2 Relevance"]
for i, h in enumerate(headers9, 1):
    ws9.cell(row=4, column=i, value=h)
style_header(ws9, 4, 5)

certs = [
    ("SOC 2 Type I", "Achieved", "All commercial plans", "trust.anthropic.com (under NDA)", "Direct — demonstrates Anthropic's control design"),
    ("SOC 2 Type II", "Achieved", "All commercial plans", "trust.anthropic.com (under NDA)", "Direct — demonstrates Anthropic's control operating effectiveness over time"),
    ("SOC 3 (Public Summary)", "Available", "All commercial plans", "Publicly available at trust.anthropic.com", "Reference in vendor risk assessment"),
    ("ISO 27001:2022", "Certified", "All commercial plans", "trust.anthropic.com", "Information Security Management — aligns with your ISO-aligned controls"),
    ("ISO/IEC 42001:2023", "Certified", "All commercial plans", "trust.anthropic.com", "AI Management Systems — demonstrates responsible AI governance"),
    ("CSA STAR Level 2", "Achieved", "All commercial plans", "CSA STAR Registry", "Cloud security assurance"),
    ("HIPAA / BAA", "Available", "Enterprise (sales-assisted only)", "Contact Anthropic sales", "Not in current SOC 2 scope (Security only)"),
]

for ri, row in enumerate(certs, 5):
    for ci, val in enumerate(row, 1):
        cell = ws9.cell(row=ri, column=ci, value=val)
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = normal_font

r = 5 + len(certs) + 2
ws9.cell(row=r, column=1, value="KEY DATA HANDLING COMMITMENTS (Both Plans)").font = subtitle_font
ws9.merge_cells(f"A{r}:E{r}")
r += 1

commitments = [
    ("No training on commercial data", "Anthropic does not use inputs/outputs from commercial plans to train models", "Commercial Terms of Service"),
    ("Data Processor role", "Anthropic acts as Data Processor; customer is Data Controller", "DPA (auto-incorporated in Commercial Terms)"),
    ("30-day default deletion", "Inputs/outputs automatically deleted within 30 days of receipt", "Privacy Policy / Data Retention docs"),
    ("Employee access restrictions", "Anthropic employees cannot access conversations by default. Access only via explicit user consent or Trust & Safety review.", "Privacy Center documentation"),
    ("Standard Contractual Clauses (SCCs)", "Included in DPA for international data transfers", "DPA"),
    ("Encryption in transit", "TLS 1.2+ for all network requests", "Trust Center / SOC 2 report"),
    ("Encryption at rest", "AES-256 encryption for stored data", "Trust Center / SOC 2 report"),
    ("Confidential inference", "TPM root of trust for model inference", "Trust Center"),
]

data_headers = ["Commitment", "Description", "Source"]
for i, h in enumerate(data_headers, 1):
    ws9.cell(row=r, column=i, value=h)
style_header(ws9, r, 3)
r += 1

for row in commitments:
    for ci, val in enumerate(row, 1):
        cell = ws9.cell(row=r, column=ci, value=val)
        cell.border = thin_border
        cell.alignment = wrap
        cell.font = normal_font
    r += 1

for i, w in enumerate([30, 35, 25, 35, 40], 1):
    ws9.column_dimensions[get_column_letter(i)].width = w


# ── Save ────────────────────────────────────────────────────────────
output_path = "/Users/evanpaliotta/Desktop/iAltA Test/Citizen AI Engineer/iAltA_SOC2_Claude_Teams_vs_Enterprise_Assessment.xlsx"
wb.save(output_path)
print(f"Workbook saved to: {output_path}")
